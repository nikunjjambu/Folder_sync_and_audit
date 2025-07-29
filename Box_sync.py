
import os
import shutil
import hashlib
import pandas as pd
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
import threading

def compute_checksum(file_path, algo='sha256'):
    hash_func = hashlib.new(algo)
    try:
        with open(file_path, 'rb') as f:
            while chunk := f.read(8192):
                hash_func.update(chunk)
        return hash_func.hexdigest()
    except Exception as e:
        return f"ERROR: {e}"

def get_file_size(file_path):
    try:
        return os.path.getsize(file_path)
    except:
        return None

def file_exists(path):
    return os.path.exists(path)

def cleanup_old_backups(report_path, keep_last=7):
    base_name = os.path.basename(report_path).replace(".xlsx", "")
    folder = os.path.dirname(report_path)
    backups = sorted([
        os.path.join(folder, f) for f in os.listdir(folder)
        if f.startswith(base_name + "_backup_") and f.endswith(".xlsx")
    ], key=os.path.getmtime, reverse=True)

    for old_backup in backups[keep_last:]:
        try:
            os.remove(old_backup)
        except Exception as e:
            print(f"Failed to delete old backup: {old_backup} - {e}")

def backup_excel(report_path):
    if os.path.exists(report_path):
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        backup_path = report_path.replace(".xlsx", f"_backup_{timestamp}.xlsx")
        shutil.copy2(report_path, backup_path)
        cleanup_old_backups(report_path, keep_last=7)

def sync_and_verify(folder1, folder2, report_path, progress_callback, status_callback, force_recopy=False):
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    error_log_entries = []
    audit_records = []

    backup_excel(report_path)

    if os.path.exists(report_path):
        df = pd.read_excel(report_path, engine='openpyxl')
    else:
        file_records = []
        for root, _, files in os.walk(folder2):
            for file in files:
                full_path = os.path.join(root, file)
                relative_path = os.path.relpath(full_path, folder2)
                file_records.append({
                    'Relative Path': relative_path,
                    'Source Path': full_path,
                    'Date Copied to Folder 1': pd.NaT,
                    'Exists in Folder 1': False,
                    'Exists in Folder 2': True
                })
        df = pd.DataFrame(file_records)

    total_files = len(df)
    copied_files = 0
    verified_files = 0
    mismatched_files = 0
    missing_files = 0

    for index, row in df.iterrows():
        relative_path = row.get('Relative Path')
        source_path = row.get('Source Path')
        dest_path = os.path.join(folder1, relative_path)

        exists1 = file_exists(dest_path)
        exists2 = file_exists(source_path)
        df.at[index, 'Exists in Folder 1'] = exists1
        df.at[index, 'Exists in Folder 2'] = exists2

        status = ""
        status_callback(f"Processing: {relative_path}")

        if not exists2:
            status = "Missing in Folder 2"
            missing_files += 1
            error_log_entries.append(f"[{datetime.now()}] Missing in Folder 2 - {relative_path}")
        elif not exists1 and force_recopy:
            df.at[index, 'Date Copied to Folder 1'] = pd.NaT

        if exists2:
            if pd.isna(df.at[index, 'Date Copied to Folder 1']):
                try:
                    os.makedirs(os.path.dirname(dest_path), exist_ok=True)
                    shutil.copy2(source_path, dest_path)
                    df.at[index, 'Date Copied to Folder 1'] = str(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                    copied_files += 1
                    status = "Copied"
                except Exception as e:
                    status = f"Error copying: {str(e)}"
                    error_log_entries.append(f"[{datetime.now()}] {status} - {relative_path}")
            else:
                status = "Already Copied"

            if file_exists(dest_path):
                source_size = get_file_size(source_path)
                dest_size = get_file_size(dest_path)

                if source_size != dest_size:
                    mismatched_files += 1
                    status = "Size Mismatch"
                else:
                    source_hash = compute_checksum(source_path)
                    dest_hash = compute_checksum(dest_path)
                    if source_hash != dest_hash:
                        mismatched_files += 1
                        status = "Checksum Mismatch"
                    else:
                        verified_files += 1
                        if status != "Copied":
                            status = "Verified"
        else:
            status = "Missing in Folder 1"
            missing_files += 1
            error_log_entries.append(f"[{datetime.now()}] Missing in Folder 1 - {relative_path}")

        audit_records.append({
            'Timestamp': timestamp,
            'Relative Path': relative_path,
            'Status': status
        })

        progress_callback(index + 1, total_files)

    df.to_excel(report_path, index=False, engine='openpyxl')

    try:
        book = load_workbook(report_path)
        audit_df = pd.DataFrame(audit_records)
        sheet_name = f"audit_{timestamp}"
        book.create_sheet(sheet_name)
        sheet = book[sheet_name]
        for r in dataframe_to_rows(audit_df, index=False, header=True):
            sheet.append(r)

        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        main_sheet = book[book.sheetnames[0]]
        for row in main_sheet.iter_rows(min_row=2, max_col=main_sheet.max_column):
            status_cell = row[main_sheet.max_column - 3]
            if status_cell.value in ["Missing in Folder 1", "Missing in Folder 2", "Size Mismatch", "Checksum Mismatch"]:
                for cell in row:
                    cell.fill = red_fill

        book.save(report_path)
    except Exception as e:
        error_log_entries.append(f"[{datetime.now()}] Failed to append audit sheet: {str(e)}")

    return copied_files, verified_files, mismatched_files, missing_files, error_log_entries

def run_sync_thread():
    run_button.config(state="disabled")
    progress_bar["value"] = 0
    status_label.config(text="Starting sync...")

    def update_progress(current, total):
        progress_bar["maximum"] = total
        progress_bar["value"] = current

    def update_status(msg):
        status_label.config(text=msg)

    def task():
        folder1 = folder1_var.get()
        folder2 = folder2_var.get()
        force_recopy = force_recopy_var.get()
        if not folder1 or not folder2:
            messagebox.showerror("Error", "Please select both folders.")
            run_button.config(state="normal")
            return

        report_path = os.path.join(os.getcwd(), "missing_files_report.xlsx")
        copied, verified, mismatched, missing, errors = sync_and_verify(
            folder1, folder2, report_path, update_progress, update_status, force_recopy
        )

        summary = (
            f"✅ Copied: {copied}\n"
            f"✅ Verified: {verified}\n"
            f"⚠️ Mismatched: {mismatched}\n"
            f"❌ Missing: {missing}\n"
        )
        if errors:
            summary += f"\n📝 Errors logged: {len(errors)}"
            with open("copy_errors.log", "w", encoding="utf-8") as f:
                f.write("\n".join(errors))

        messagebox.showinfo("Sync Summary", summary)
        run_button.config(state="normal")
        status_label.config(text="Sync complete.")

    threading.Thread(target=task).start()

def browse_folder1():
    path = filedialog.askdirectory()
    if path:
        folder1_var.set(path)

def browse_folder2():
    path = filedialog.askdirectory()
    if path:
        folder2_var.set(path)

root = tk.Tk()
root.resizable(False, False)
root.title("Box Folder Sync & Audit")

frame = ttk.Frame(root, padding=20)
frame.grid()

folder1_var = tk.StringVar()
folder2_var = tk.StringVar()
force_recopy_var = tk.BooleanVar()

ttk.Label(frame, text="Internal Folder (Folder 1):").grid(column=0, row=0, sticky="w")
ttk.Entry(frame, textvariable=folder1_var, width=60).grid(column=1, row=0)
ttk.Button(frame, text="Browse", command=browse_folder1).grid(column=2, row=0)

ttk.Label(frame, text="External Folder (Folder 2):").grid(column=0, row=1, sticky="w")
ttk.Entry(frame, textvariable=folder2_var, width=60).grid(column=1, row=1)
ttk.Button(frame, text="Browse", command=browse_folder2).grid(column=2, row=1)

ttk.Checkbutton(frame, text="Force re-copy files missing in Folder 1", variable=force_recopy_var).grid(column=1, row=2, sticky="w")

run_button = ttk.Button(frame, text="Run Sync & Verify", command=run_sync_thread)
run_button.grid(column=1, row=3, pady=10)

progress_bar = ttk.Progressbar(frame, orient="horizontal", length=400, mode="determinate")
progress_bar.grid(column=1, row=4, pady=5)

status_label = ttk.Label(frame, text="Ready.", width=60)
status_label.grid(column=1, row=5, pady=5)

root.mainloop()
